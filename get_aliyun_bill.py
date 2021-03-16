#!/usr/bin/env python
#coding=utf-8
#需要安装openpyxl

from aliyunsdkcore.client import AcsClient
from aliyunsdkcore.acs_exception.exceptions import ClientException
from aliyunsdkcore.acs_exception.exceptions import ServerException
from aliyunsdkbssopenapi.request.v20171214.QueryAccountBalanceRequest import QueryAccountBalanceRequest
from aliyunsdkbssopenapi.request.v20171214.QueryAccountBillRequest import QueryAccountBillRequest
from aliyunsdkbssopenapi.request.v20171214.QueryBillOverviewRequest import QueryBillOverviewRequest

import os
import sys
import json
import openpyxl
import datetime
import string

AccessKeyID = "xxxxxxxxxxxxxxxxxxxxxxxxxxxxx"
AccessKeySecret = "xxxxxxxxxxxxxxxxxxxxxxxxxxxxxx"

# 阿里云各类型账单
class GetAliBilling():

    def __init__(self):
        self.client = AcsClient(AccessKeyID, AccessKeySecret, 'cn-hangzhou')

    # 账户余额
    def getBalance(self):
        request = QueryAccountBalanceRequest()
        request.set_accept_format('json')
        response = self.client.do_action_with_exception(request)
        #print(response) 
        return response
        #print("-----------------------------------------------")

    # 月总金额
    def getMonthTotalBill(self, month):
        request = QueryAccountBillRequest()
        request.set_accept_format('json')
        try:
            request.set_BillingCycle(month)
        except Exception as e:
            print("日期格式不对%s" % str(e))
        response = self.client.do_action_with_exception(request)
        #print(response)
        return response 
        #print("-----------------------------------------------")

    # 月产品明细金额
    def getMonthDetail(self, month):
        request = QueryBillOverviewRequest()
        request.set_accept_format('json')
        try:
            request.set_BillingCycle(month)
        except Exception as e:
            print("日期格式不对%s" % str(e))
        response = self.client.do_action_with_exception(request)
        #print(response) 
        return response
        #print("-----------------------------------------------")

# excel表格操作
class operExcel():

    def __init__(self):
        self.al_list = list(string.ascii_lowercase)

    # 生成表格写入数据
    def writeExcel(self, data, timer):
        date = data["Data"]["BillingCycle"]
        cost_detail = data["Data"]["Items"]["Item"]
        menu = ["产品明细", "实际金额"]
        cost_name = []
        cost_original = []
        cost_num = []
        file_name = os.getcwd() + "/" + "%s阿里云月费用明细.xlsx" % timer
        for item in cost_detail:
            amount = item["PaymentAmount"]
            num_list = []
            for i in range(len(cost_detail) - 1):
                if item["ProductDetail"] == cost_detail[i]["ProductDetail"] and item["PaymentAmount"] != cost_detail[i]["PaymentAmount"]:
                    amount = amount + cost_detail[i]["PaymentAmount"]
                    num_list.append(i)
                else:
                    amount = amount
            cost_name.append(item["ProductDetail"])
            cost_num.append(amount)
            # cost_original.append(item["PretaxGrossAmount"])
            # cost_num.append(item["PaymentAmount"])
            for n in reversed(num_list):
                del cost_detail[n]
        if not os.path.exists(file_name): # 文件不存在
            wb = openpyxl.Workbook() # 创建表格
            ws = wb.active # 获取当前活跃的worksheet,默认就是第一个worksheet
            ws.title = u'费用明细' # 设置表名
            ws.merge_cells("A1:B1") # 合并单元格
            ws.cell(1, 1, date) # 插入数据--时间
            for i in range(len(menu)):
                ws.cell(2, i + 1, menu[i])
            for i in range(len(cost_detail)):
                ws.cell(i + 3, 1, cost_name[i])
                #ws.cell(i + 3, 2, cost_original[i])
                ws.cell(i + 3, 2, cost_num[i])
        else: # 文件存在
            wb = openpyxl.load_workbook(file_name) # 从文件中读取
            ws = wb.active # 获取当前活跃的worksheet,默认就是第一个worksheet
            ws.insert_cols(1) # 插入第一行
            ws.insert_cols(1) # 插入第一行
            #ws.merge_cells("C1:D1") # 合并单元格
            ws.cell(1, 1, date) # 插入数据--时间
            for i in range(len(menu)):
                ws.cell(2, i + 1, menu[i])
            # 按上月对应产品明细排列
            new_item = []
            tmp_item = []
            last_acount = []
            for row in ws.iter_rows(min_row=3, min_col=3, max_col=3, max_row=int(ws.max_row) - 1):
                for cell in row:
                    for n in range(len(cost_name)):
                        if cost_name[n] == cell.value:
                            ws.cell(cell.row, 1, cost_name[n])
                            ws.cell(cell.row, 2, cost_num[n])
                            new_item.append(cost_name[n])
                            continue
            for i in range(len(cost_name)):
                if cost_name[i] not in new_item:
                    tmp_item.append({"ProductDetail": cost_name[i], "PaymentAmount": cost_num[i]})
            
            for row in ws.iter_rows(min_row=3, min_col=1, max_col=1):
                if len(tmp_item) > 0:
                    for cell in row:
                        if not cell.value:
                            ws.cell(cell.row, 1, tmp_item[0]["ProductDetail"])
                            ws.cell(cell.row, 2, tmp_item[0]["PaymentAmount"])
                            del tmp_item[0]
                            break
                else:
                    break
            for i in range(3, ws.max_row): # 对比数据
                result = ws.cell(i, 2).value - ws.cell(i, 4).value
                ws.cell(i, 5, result)
            ws.cell(2, 5, value="月对比金额")
            ws.cell(1, 5).fill = openpyxl.styles.PatternFill("solid", fgColor="00339966")
            ws.cell(2, 5).fill = openpyxl.styles.PatternFill("solid", fgColor="00339966")
            for row in tuple(ws["E" + str(1) : "E" + str(ws.max_row)]):
                for cell in row:
                    cell.border = self.border('thin', 'thin', 'thin', 'thin')
        max_row = ws.max_row
        # 总计
        acount = float(0)
        for row in ws.iter_rows(min_row=3, min_col=2, max_col=2, max_row=max_row):
            for cell in row:
                acount = acount + float(cell.value)
        ws.cell(max_row, 1, "总计")
        ws.cell(max_row, 2, acount)
        #设置列宽
        max_col = ws.max_column / 2
        for i in range(max_col):
            ws.column_dimensions[self.al_list[i * 2]].width=25 # 设置列宽
            ws.merge_cells(self.al_list[i * 2] + str(1) + ":" + self.al_list[(i * 2) + 1] + str(1)) # 合并单元格
        # 设置颜色,字体
        for row in tuple(ws["A" + str(1) : "B" + str(2)]):
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor="00339966")
                cell.font = self.Font(bold=True)
        # 设置指定区域边框
        for row in tuple(ws["A" + str(1) : "B" + str(max_row)]):
            for cell in row:
                cell.border = self.border('thin', 'thin', 'thin', 'thin')
        wb.save('%s阿里云月费用明细.xlsx' % timer)

    # 设置边框
    def border(self,left, right, top, bottom):
        border = openpyxl.styles.Border(left=openpyxl.styles.Side(style=left, color='000000'),
                                        right=openpyxl.styles.Side(style=right, color='000000'),
                                        top=openpyxl.styles.Side(style=top,color='000000'),
                                        bottom=openpyxl.styles.Side(style=bottom, color='000000'))
        return border
    def Font(self, name="Arial", size="11", bold=False):
        font = openpyxl.styles.Font(name=name, size=size, bold=bold)

if __name__ == "__main__":
    #last_two_month = str(datetime.datetime.now().year) + "-" + str(datetime.datetime.now().month - 2) #last two month
    last_month = str(datetime.datetime.now().year) + "-" + str(datetime.datetime.now().month - 1) #last month
    if len(sys.argv) == 2:
        date_t = sys.argv[1]
    else:
        date_t = last_month
    bill = GetAliBilling()
    #bill.getBalance()
    total = json.loads(bill.getMonthTotalBill(date_t))["Data"]["Items"]["Item"][0]["PaymentAmount"]
    data = json.loads(bill.getMonthDetail(date_t))
    excel = operExcel()
    excel.writeExcel(data, last_month)
