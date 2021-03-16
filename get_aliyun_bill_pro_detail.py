#!/usr/bin/env python
#coding=utf-8
# explain：从阿里云导出账单明细，根据账单明细计算汇总各产品明细，需要接收一个参数，参数为excel表名
# 导出阿里云账单：费用-->费用账单-->账单明细-->统计项（费用），统计周期（按天）-->导出账单csv

import openpyxl
import json
import sys
import datetime
import os

class statisticsAliyunBill():

    def __init__(self, file_name):
        self.name = file_name

    # 根据现有表格生成数据
    def statistics(self):
        wb = openpyxl.load_workbook(self.name) # 从文件中读取
        ws = wb.active # 获取当前活跃的worksheet,默认就是第一个worksheet 
        date_time = ws.cell(row=2, column=1).value # 获取日期
        product_name = []
        product_list = []
        acount_list = []
        
        for row in ws.iter_rows(max_col=11, min_col=11,min_row=2, max_row=ws.max_row): # 读取11列按产品明细统计
            for cell in row: 
                product_name.append(cell.value)
        product_name = set(product_name)
        product_name = list(product_name)
        for name in product_name:
            ids = []
            for row in ws.iter_rows(max_col=11, min_col=11,min_row=2, max_row=ws.max_row): 
                for cell in row: 
                    if cell.value == name:
                        ids.append(ws.cell(row=cell.row, column=17).value)  # 产品明细--实例ID
            ids = set(ids)
            ids = list(ids)
            ids.sort()
            product_list.append({name: ids})  # [{"name":[xx,xx,xx]}, {"name1":[xx,xx,xx]}]

        # 按产品id遍历产品明细，累计每个产品id价格
        for item in product_list:
            for name in item:
                id_acount = []
                for n in item[name]:
                    acount = float(0)
                    for row in ws.iter_rows(max_col=17, min_col=17,min_row=2, max_row=ws.max_row):
                        for cell in row: 
                            if cell.value == n:
                                acount = acount + ws.cell(row=cell.row, column=38).value
                    id_acount.append([n, acount])
                acount_list.append({name: id_acount})
        return acount_list, date_time

    # 写入excel操作
    def writeExcel(self, data, date_time):
        #last_month = str(datetime.datetime.now().year) + "-" + str(datetime.datetime.now().month - 1) #last month
        file_name = os.getcwd() + "/" + "%s阿里云产品明细费用统计.xlsx" % str(date_time.strip())
        if not os.path.exists(file_name): # 文件不存在
            wb = openpyxl.Workbook() # 创建表格
            for item in data:
                for name in item:
                    ws = wb.create_sheet(name)  # 创建工作表
                    #sheet = wb[name]
                    sheet = wb.get_sheet_by_name(name)  # 选择工作表
                    sheet.cell(row=1, column=1, value=date_time)
                    for i in item[name]:  # 写入数据
                        sheet.append(i)
            del wb["Sheet"]  # 删除默认Sheet工作表
        else:
            wb = openpyxl.load_workbook(file_name) # 从文件中读取
            for item in data:
                for name in item:
                    #sheet = wb[name]
                    sheet = wb.get_sheet_by_name(name)  # 选择工作表
                    sheet.insert_cols(1) # 插入第一行
                    sheet.insert_cols(1) # 插入第一行
                    sheet.cell(row=1, column=1, value=date_time)
                    for i in range(len(item[name])):  # 写入数据
                        sheet.cell(i + 2, 1, item[name][i][0])
                        sheet.cell(i + 2, 2, item[name][i][1])
        wb.save(file_name)


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("USAGE: python %s file_name.xlsx" % sys.argv[0])
        sys.exit()
    file_name = sys.argv[1]
    Excel = statisticsAliyunBill(file_name)
    data, date_time = Excel.statistics()
    Excel.writeExcel(data, date_time)
